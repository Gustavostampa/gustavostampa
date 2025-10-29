from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import hashlib
import openpyxl
from io import BytesIO
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class Usuario(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    login: str
    senha_hash: str
    perfil: str  # gestor ou conferente
    ativo: bool = True
    criado_em: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UsuarioCreate(BaseModel):
    nome: str
    login: str
    senha: str
    perfil: str

class LoginRequest(BaseModel):
    login: str
    senha: str

class Produto(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    codigo_produto: str
    descricao: str
    ean: str
    tipo_unidade: str
    ativo: bool = True
    criado_em: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ProdutoEAN(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku: str  # codigo_produto
    ean: str  # EAN normalizado
    tipo_unidade: str  # 'UNI', 'CX', 'EXB', 'FRD', etc
    descricao: Optional[str] = None  # descrição específica desta embalagem (ex: "Caixa com 12 unidades")
    ativo: bool = True
    criado_em: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ItemCarga(BaseModel):
    codigo_produto: str
    descricao: str
    unidade: str
    quantidade: int
    quantidade_conferida: int = 0
    status: str = "pendente"  # pendente, ok, diferenca, fora_lista
    recipiente: Optional[str] = None
    ean: Optional[str] = None

class Carga(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    identificador_carga: str
    data: str
    tipo: str  # caixaria ou multi
    status: str = "aguardando"  # aguardando, em_andamento, pausada, finalizada
    itens: List[ItemCarga]
    conferente_id: Optional[str] = None
    criado_em: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Sessao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    carga_id: str
    conferente_id: str
    recipiente: Optional[str] = None  # Para Multi-pedidos
    inicio: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    pausas: List[Dict[str, str]] = []
    fim: Optional[str] = None
    status: str = "ativa"  # ativa, pausada, finalizada

class Sobra(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    carga_id: str
    sessao_id: str
    recipiente: Optional[str] = None  # Para Multi
    ean: str
    descricao: Optional[str] = None
    quantidade: int = 0
    primeira_leitura: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    ultima_leitura: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    observacao: Optional[str] = None

class Leitura(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sessao_id: str
    carga_id: str
    conferente_id: str
    ean: str
    quantidade: int
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    resultado: str  # ok, diferenca, fora_lista

class LeituraCreate(BaseModel):
    sessao_id: str
    carga_id: str
    ean: str
    quantidade: int = 1

class SessaoCreate(BaseModel):
    carga_id: str
    recipiente: Optional[str] = None  # Obrigatório para Multi-pedidos

def hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode()).hexdigest()

def normalizar_ean(ean: str) -> str:
    """
    Normaliza EAN removendo espaços, hífens e validando formato.
    Aceita GTIN-8, GTIN-12, GTIN-13, GTIN-14.
    Retorna string numérica padronizada.
    """
    if not ean:
        return ""
    
    # Remove espaços, hífens e caracteres não numéricos
    ean_limpo = ''.join(c for c in ean if c.isdigit())
    
    # Valida comprimento (EAN-8, EAN-13, UPC-12, GTIN-14)
    if len(ean_limpo) not in [8, 12, 13, 14]:
        return ean_limpo  # Retorna mesmo assim para permitir EANs customizados
    
    # Pad com zeros à esquerda para garantir comprimento consistente
    # Armazena sempre como 14 dígitos (GTIN-14 é o padrão mais amplo)
    return ean_limpo.zfill(14)

# Auth
@api_router.post("/auth/login")
async def login(request: LoginRequest):
    senha_hash = hash_senha(request.senha)
    usuario = await db.usuarios.find_one(
        {"login": request.login, "senha_hash": senha_hash, "ativo": True},
        {"_id": 0}
    )
    if not usuario:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    return usuario

# Usuários
@api_router.get("/usuarios", response_model=List[Usuario])
async def listar_usuarios():
    usuarios = await db.usuarios.find({}, {"_id": 0}).to_list(1000)
    return usuarios

@api_router.post("/usuarios", response_model=Usuario)
async def criar_usuario(input: UsuarioCreate):
    # Verificar se login já existe
    existe = await db.usuarios.find_one({"login": input.login})
    if existe:
        raise HTTPException(status_code=400, detail="Login já existe")
    
    usuario = Usuario(
        nome=input.nome,
        login=input.login,
        senha_hash=hash_senha(input.senha),
        perfil=input.perfil
    )
    await db.usuarios.insert_one(usuario.model_dump())
    return usuario

@api_router.delete("/usuarios/{usuario_id}")
async def deletar_usuario(usuario_id: str):
    result = await db.usuarios.update_one(
        {"id": usuario_id},
        {"$set": {"ativo": False}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return {"message": "Usuário desativado com sucesso"}

# Produtos
@api_router.get("/produtos", response_model=List[Produto])
async def listar_produtos(
    descricao: Optional[str] = None,
    codigo_produto: Optional[str] = None,
    ean: Optional[str] = None,
    page: int = 1,
    limit: int = 50
):
    filtro = {}
    if descricao:
        filtro["descricao"] = {"$regex": descricao, "$options": "i"}
    if codigo_produto:
        filtro["codigo_produto"] = codigo_produto
    if ean:
        filtro["ean"] = ean
    
    skip = (page - 1) * limit
    produtos = await db.produtos.find(filtro, {"_id": 0}).sort("descricao", 1).skip(skip).limit(limit).to_list(limit)
    return produtos

@api_router.get("/produtos/count")
async def contar_produtos(
    descricao: Optional[str] = None,
    codigo_produto: Optional[str] = None,
    ean: Optional[str] = None
):
    filtro = {}
    if descricao:
        filtro["descricao"] = {"$regex": descricao, "$options": "i"}
    if codigo_produto:
        filtro["codigo_produto"] = codigo_produto
    if ean:
        filtro["ean"] = ean
    
    total = await db.produtos.count_documents(filtro)
    return {"total": total}

class ProdutoCreate(BaseModel):
    codigo_produto: str
    descricao: str
    ean: str
    tipo_unidade: str
    ativo: bool = True

class ProdutoUpdate(BaseModel):
    codigo_produto: Optional[str] = None
    descricao: Optional[str] = None
    ean: Optional[str] = None
    tipo_unidade: Optional[str] = None
    ativo: Optional[bool] = None

@api_router.post("/produtos/criar", response_model=Produto)
async def criar_produto(input: ProdutoCreate):
    # Validar tipo_unidade
    if input.tipo_unidade not in ["UNI", "CX", "EXB"]:
        raise HTTPException(status_code=400, detail="tipo_unidade deve ser UNI, CX ou EXB")
    
    # Verificar se EAN já existe
    existe = await db.produtos.find_one({"ean": input.ean})
    if existe:
        raise HTTPException(status_code=400, detail="EAN já cadastrado")
    
    produto = Produto(**input.model_dump())
    await db.produtos.insert_one(produto.model_dump())
    return produto

@api_router.put("/produtos/{produto_id}", response_model=Produto)
async def atualizar_produto(produto_id: str, input: ProdutoUpdate):
    # Buscar produto
    produto = await db.produtos.find_one({"id": produto_id})
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    
    # Preparar atualização
    update_data = {}
    if input.codigo_produto is not None:
        update_data["codigo_produto"] = input.codigo_produto
    if input.descricao is not None:
        update_data["descricao"] = input.descricao
    if input.ean is not None:
        # Verificar se novo EAN já existe
        if input.ean != produto["ean"]:
            existe = await db.produtos.find_one({"ean": input.ean})
            if existe:
                raise HTTPException(status_code=400, detail="EAN já cadastrado")
        update_data["ean"] = input.ean
    if input.tipo_unidade is not None:
        if input.tipo_unidade not in ["UNI", "CX", "EXB"]:
            raise HTTPException(status_code=400, detail="tipo_unidade deve ser UNI, CX ou EXB")
        update_data["tipo_unidade"] = input.tipo_unidade
    if input.ativo is not None:
        update_data["ativo"] = input.ativo
    
    await db.produtos.update_one({"id": produto_id}, {"$set": update_data})
    produto_atualizado = await db.produtos.find_one({"id": produto_id}, {"_id": 0})
    return produto_atualizado

@api_router.delete("/produtos/{produto_id}")
async def deletar_produto(produto_id: str):
    # Buscar produto
    produto = await db.produtos.find_one({"id": produto_id})
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    
    # Verificar se está em cargas não finalizadas
    cargas_ativas = await db.cargas.find_one({
        "status": {"$ne": "finalizada"},
        "itens.codigo_produto": produto["codigo_produto"]
    })
    
    if cargas_ativas:
        raise HTTPException(
            status_code=400, 
            detail="Produto está em cargas não finalizadas. Use desativar ao invés de excluir."
        )
    
    await db.produtos.delete_one({"id": produto_id})
    return {"message": "Produto excluído com sucesso"}

# Produto EANs
@api_router.get("/produto-eans", response_model=List[ProdutoEAN])
async def listar_produto_eans(sku: Optional[str] = None, ean: Optional[str] = None):
    filtro = {}
    if sku:
        filtro["sku"] = sku
    if ean:
        ean_norm = normalizar_ean(ean)
        filtro["ean"] = ean_norm
    
    eans = await db.produto_eans.find(filtro, {"_id": 0}).to_list(1000)
    return eans

@api_router.get("/produto-eans/buscar-por-ean/{ean}")
async def buscar_produto_ean_por_ean(ean: str):
    ean_norm = normalizar_ean(ean)
    produto_ean = await db.produto_eans.find_one({"ean": ean_norm, "ativo": True}, {"_id": 0})
    if not produto_ean:
        return None
    return produto_ean

@api_router.post("/produto-eans/criar", response_model=ProdutoEAN)
async def criar_produto_ean(produto_ean: ProdutoEAN):
    # Normalizar EAN
    produto_ean.ean = normalizar_ean(produto_ean.ean)
    
    # Verificar se já existe este EAN
    existe = await db.produto_eans.find_one({"ean": produto_ean.ean})
    if existe:
        raise HTTPException(status_code=400, detail="EAN já cadastrado")
    
    # Verificar se SKU existe
    produto = await db.produtos.find_one({"codigo_produto": produto_ean.sku})
    if not produto:
        raise HTTPException(status_code=404, detail="SKU não encontrado no cadastro de produtos")
    
    await db.produto_eans.insert_one(produto_ean.model_dump())
    return produto_ean

@api_router.put("/produto-eans/{produto_ean_id}", response_model=ProdutoEAN)
async def atualizar_produto_ean(produto_ean_id: str, update_data: Dict[str, Any]):
    # Normalizar EAN se fornecido
    if "ean" in update_data:
        update_data["ean"] = normalizar_ean(update_data["ean"])
        
        # Verificar duplicidade
        existe = await db.produto_eans.find_one({
            "ean": update_data["ean"],
            "id": {"$ne": produto_ean_id}
        })
        if existe:
            raise HTTPException(status_code=400, detail="EAN já cadastrado")
    
    result = await db.produto_eans.update_one(
        {"id": produto_ean_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produto EAN não encontrado")
    
    updated = await db.produto_eans.find_one({"id": produto_ean_id}, {"_id": 0})
    return ProdutoEAN(**updated)

@api_router.delete("/produto-eans/{produto_ean_id}")
async def deletar_produto_ean(produto_ean_id: str):
    result = await db.produto_eans.delete_one({"id": produto_ean_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Produto EAN não encontrado")
    return {"message": "Produto EAN excluído com sucesso"}

# Importações
@api_router.post("/importar/produtos")
async def importar_produtos(file: UploadFile = File(...), acao: str = Form("substituir")):
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    preview = []
    produtos_validos = []
    erros = []
    duplicados = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if idx > 51:  # Preview 50 linhas + header
            break
        
        if not row[0]:  # Skip empty rows
            continue
            
        codigo_produto, descricao, ean, tipo_unidade = row[:4]
        
        # Validação
        if not all([codigo_produto, descricao, ean, tipo_unidade]):
            erros.append(f"Linha {idx}: campos obrigatórios faltando")
            continue
        
        if tipo_unidade not in ["UNI", "CX", "EXB"]:
            erros.append(f"Linha {idx}: tipo_unidade deve ser UNI, CX ou EXB")
            continue
        
        produto = {
            "codigo_produto": str(codigo_produto),
            "descricao": str(descricao),
            "ean": str(ean),
            "tipo_unidade": str(tipo_unidade)
        }
        
        # Verificar duplicidade
        existe = await db.produtos.find_one({"ean": str(ean)})
        if existe:
            duplicados.append(str(ean))
            if acao == "ignorar":
                continue
        
        produtos_validos.append(produto)
        if idx <= 51:
            preview.append(produto)
    
    # Processar importação completa
    if acao == "substituir" and produtos_validos:
        for p in produtos_validos:
            await db.produtos.update_one(
                {"ean": p["ean"]},
                {"$set": {**p, "id": str(uuid.uuid4()), "criado_em": datetime.now(timezone.utc).isoformat()}},
                upsert=True
            )
    elif acao != "substituir" and produtos_validos:
        # Inserir apenas novos
        novos = [p for p in produtos_validos if p["ean"] not in duplicados]
        if novos:
            docs = [{**p, "id": str(uuid.uuid4()), "criado_em": datetime.now(timezone.utc).isoformat()} for p in novos]
            await db.produtos.insert_many(docs)
    
    return {
        "preview": preview,
        "total_processados": len(produtos_validos),
        "erros": erros,
        "duplicados": duplicados
    }

@api_router.post("/importar/caixaria")
async def importar_caixaria(file: UploadFile = File(...), acao: str = Form("substituir")):
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    cargas_dict = {}
    preview = []
    erros = []
    duplicados = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        identificador_carga, data, codigo_produto, descricao, unidade, quantidade = row[:6]
        
        if not all([identificador_carga, data, codigo_produto, descricao, unidade, quantidade]):
            erros.append(f"Linha {idx}: campos obrigatórios faltando")
            continue
        
        # Normalizar data (remover timestamp se existir)
        data_str = str(data)
        if ' ' in data_str:
            data_str = data_str.split(' ')[0]
        
        if str(identificador_carga) not in cargas_dict:
            cargas_dict[str(identificador_carga)] = {
                "identificador_carga": str(identificador_carga),
                "data": data_str,
                "tipo": "caixaria",
                "itens": []
            }
        
        item = ItemCarga(
            codigo_produto=str(codigo_produto),
            descricao=str(descricao),
            unidade=str(unidade),
            quantidade=int(quantidade)
        )
        cargas_dict[str(identificador_carga)]["itens"].append(item.model_dump())
        
        if len(preview) < 50:
            preview.append({"identificador_carga": str(identificador_carga), **item.model_dump()})
    
    # Salvar cargas
    for carga_data in cargas_dict.values():
        # Buscar EANs dos produtos
        for item in carga_data["itens"]:
            produto = await db.produtos.find_one({"codigo_produto": item["codigo_produto"]})
            if produto:
                item["ean"] = produto["ean"]
        
        existe = await db.cargas.find_one({"identificador_carga": carga_data["identificador_carga"]})
        if existe:
            duplicados.append(carga_data["identificador_carga"])
            if acao == "substituir":
                await db.cargas.delete_one({"identificador_carga": carga_data["identificador_carga"]})
            else:
                continue
        
        carga = Carga(**carga_data)
        await db.cargas.insert_one(carga.model_dump())
    
    return {
        "preview": preview,
        "total_cargas": len(cargas_dict),
        "erros": erros,
        "duplicados": duplicados
    }

@api_router.post("/importar/multi")
async def importar_multi(file: UploadFile = File(...), acao: str = Form("substituir")):
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    cargas_dict = {}
    preview = []
    erros = []
    duplicados = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        identificador_carga, data, recipiente, codigo_produto, descricao, unidade, quantidade = row[:7]
        
        if not all([identificador_carga, data, recipiente, codigo_produto, descricao, unidade, quantidade]):
            erros.append(f"Linha {idx}: campos obrigatórios faltando")
            continue
        
        # Normalizar data (remover timestamp se existir)
        data_str = str(data)
        if ' ' in data_str:
            data_str = data_str.split(' ')[0]
        
        if str(identificador_carga) not in cargas_dict:
            cargas_dict[str(identificador_carga)] = {
                "identificador_carga": str(identificador_carga),
                "data": data_str,
                "tipo": "multi",
                "itens": []
            }
        
        item = ItemCarga(
            codigo_produto=str(codigo_produto),
            descricao=str(descricao),
            unidade=str(unidade),
            quantidade=int(quantidade),
            recipiente=str(recipiente)
        )
        cargas_dict[str(identificador_carga)]["itens"].append(item.model_dump())
        
        if len(preview) < 50:
            preview.append({"identificador_carga": str(identificador_carga), "recipiente": str(recipiente), **item.model_dump()})
    
    # Salvar cargas
    for carga_data in cargas_dict.values():
        # Buscar EANs dos produtos
        for item in carga_data["itens"]:
            produto = await db.produtos.find_one({"codigo_produto": item["codigo_produto"]})
            if produto:
                item["ean"] = produto["ean"]
        
        existe = await db.cargas.find_one({"identificador_carga": carga_data["identificador_carga"]})
        if existe:
            duplicados.append(carga_data["identificador_carga"])
            if acao == "substituir":
                await db.cargas.delete_one({"identificador_carga": carga_data["identificador_carga"]})
            else:
                continue
        
        carga = Carga(**carga_data)
        await db.cargas.insert_one(carga.model_dump())
    
    return {
        "preview": preview,
        "total_cargas": len(cargas_dict),
        "erros": erros,
        "duplicados": duplicados
    }

# Cargas
@api_router.get("/cargas", response_model=List[Carga])
async def listar_cargas(data: Optional[str] = None, tipo: Optional[str] = None, status: Optional[str] = None):
    filtro = {}
    if data:
        filtro["data"] = data
    if tipo:
        filtro["tipo"] = tipo
    if status:
        filtro["status"] = status
    
    cargas = await db.cargas.find(filtro, {"_id": 0}).to_list(1000)
    return cargas

@api_router.get("/cargas/ultima-atualizacao")
async def obter_ultima_atualizacao():
    # Retorna timestamp da última modificação em cargas
    ultima = await db.cargas.find_one({}, {"_id": 0, "criado_em": 1}, sort=[("criado_em", -1)])
    if ultima:
        return {"ultima_atualizacao": ultima.get("criado_em", datetime.now(timezone.utc).isoformat())}
    return {"ultima_atualizacao": datetime.now(timezone.utc).isoformat()}

@api_router.get("/cargas/{carga_id}", response_model=Carga)
async def obter_carga(carga_id: str):
    carga = await db.cargas.find_one({"id": carga_id}, {"_id": 0})
    if not carga:
        raise HTTPException(status_code=404, detail="Carga não encontrada")
    return carga

class ItemUpdate(BaseModel):
    quantidade_conferida: int
    status: str

@api_router.put("/cargas/{carga_id}/item/{item_index}")
async def atualizar_item_carga(carga_id: str, item_index: int, update: ItemUpdate):
    # Buscar carga
    carga = await db.cargas.find_one({"id": carga_id})
    if not carga:
        raise HTTPException(status_code=404, detail="Carga não encontrada")
    
    # Validar índice
    if item_index < 0 or item_index >= len(carga["itens"]):
        raise HTTPException(status_code=400, detail="Índice de item inválido")
    
    # Atualizar item
    carga["itens"][item_index]["quantidade_conferida"] = update.quantidade_conferida
    carga["itens"][item_index]["status"] = update.status
    
    # Salvar no banco
    await db.cargas.update_one(
        {"id": carga_id},
        {"$set": {"itens": carga["itens"]}}
    )
    
    return {"message": "Item atualizado com sucesso"}

# Sessões
@api_router.post("/sessoes", response_model=Sessao)
async def criar_sessao(input: SessaoCreate, conferente_id: str):
    # Verificar se conferente já tem sessão ativa
    sessao_ativa = await db.sessoes.find_one({
        "conferente_id": conferente_id,
        "status": {" $in": ["ativa", "pausada"]}
    })
    if sessao_ativa and sessao_ativa["status"] == "ativa":
        raise HTTPException(status_code=400, detail="Conferente já possui uma sessão ativa")
    
    # Buscar carga para validação
    carga = await db.cargas.find_one({"id": input.carga_id})
    if not carga:
        raise HTTPException(status_code=404, detail="Carga não encontrada")
    
    # Se for Multi-pedidos, recipiente é obrigatório
    if carga["tipo"] == "multi":
        if not input.recipiente:
            raise HTTPException(status_code=400, detail="Recipiente é obrigatório para Multi-pedidos")
        
        # Verificar se recipiente existe na carga
        recipientes = set(item["recipiente"] for item in carga["itens"] if item.get("recipiente"))
        if input.recipiente not in recipientes:
            raise HTTPException(status_code=400, detail="Recipiente não encontrado nesta carga. Selecione um recipiente válido.")
    
    # Atualizar status da carga
    await db.cargas.update_one(
        {"id": input.carga_id},
        {"$set": {"status": "em_andamento", "conferente_id": conferente_id}}
    )
    
    sessao = Sessao(
        carga_id=input.carga_id, 
        conferente_id=conferente_id,
        recipiente=input.recipiente
    )
    await db.sessoes.insert_one(sessao.model_dump())
    return sessao

@api_router.post("/sessoes/{sessao_id}/pausar")
async def pausar_sessao(sessao_id: str, conferente_id: str):
    # Verificar se já existe outra sessão pausada
    outra_pausada = await db.sessoes.find_one({
        "conferente_id": conferente_id,
        "status": "pausada",
        "id": {"$ne": sessao_id}
    })
    
    if outra_pausada:
        raise HTTPException(
            status_code=400, 
            detail="Você já possui uma carga pausada. Retome ou finalize a carga pausada antes de pausar outra."
        )
    
    sessao = await db.sessoes.find_one({"id": sessao_id})
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    
    pausas = sessao.get("pausas", [])
    pausas.append({"inicio": datetime.now(timezone.utc).isoformat()})
    
    await db.sessoes.update_one(
        {"id": sessao_id},
        {"$set": {"status": "pausada", "pausas": pausas}}
    )
    
    await db.cargas.update_one(
        {"id": sessao["carga_id"]},
        {"$set": {"status": "pausada"}}
    )
    
    return {"message": "Sessão pausada"}

@api_router.post("/sessoes/{sessao_id}/retomar")
async def retomar_sessao(sessao_id: str):
    sessao = await db.sessoes.find_one({"id": sessao_id})
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    
    pausas = sessao.get("pausas", [])
    if pausas and not pausas[-1].get("fim"):
        pausas[-1]["fim"] = datetime.now(timezone.utc).isoformat()
    
    await db.sessoes.update_one(
        {"id": sessao_id},
        {"$set": {"status": "ativa", "pausas": pausas}}
    )
    
    await db.cargas.update_one(
        {"id": sessao["carga_id"]},
        {"$set": {"status": "em_andamento"}}
    )
    
    return {"message": "Sessão retomada"}

@api_router.post("/sessoes/{sessao_id}/finalizar")
async def finalizar_sessao(sessao_id: str):
    sessao = await db.sessoes.find_one({"id": sessao_id})
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    
    await db.sessoes.update_one(
        {"id": sessao_id},
        {"$set": {"status": "finalizada", "fim": datetime.now(timezone.utc).isoformat()}}
    )
    
    await db.cargas.update_one(
        {"id": sessao["carga_id"]},
        {"$set": {"status": "finalizada"}}
    )
    
    return {"message": "Sessão finalizada"}

@api_router.get("/sessoes/ativa/{conferente_id}", response_model=Optional[Sessao])
async def obter_sessao_ativa(conferente_id: str):
    sessao = await db.sessoes.find_one(
        {"conferente_id": conferente_id, "status": {"$in": ["ativa", "pausada"]}},
        {"_id": 0}
    )
    return sessao

@api_router.get("/sessoes/{sessao_id}", response_model=Sessao)
async def obter_sessao_por_id(sessao_id: str):
    sessao = await db.sessoes.find_one({"id": sessao_id}, {"_id": 0})
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    return sessao

@api_router.get("/sessoes", response_model=List[Sessao])
async def listar_sessoes(
    status: Optional[str] = None,
    conferente_id: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    limit: int = 100,
    skip: int = 0
):
    filtro = {}
    
    if status:
        filtro["status"] = status
    
    if conferente_id:
        filtro["conferente_id"] = conferente_id
    
    # Filtro por data de início
    if data_inicio or data_fim:
        filtro["inicio"] = {}
        if data_inicio:
            filtro["inicio"]["$gte"] = data_inicio
        if data_fim:
            # Adicionar 1 dia para incluir todo o dia final
            dt_fim = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
            dt_fim_plus = (dt_fim + timedelta(days=1)).isoformat()
            filtro["inicio"]["$lt"] = dt_fim_plus
    
    sessoes = await db.sessoes.find(
        filtro,
        {"_id": 0}
    ).sort("inicio", -1).skip(skip).limit(limit).to_list(limit)
    
    return sessoes

# Leituras
@api_router.post("/leituras", response_model=Leitura)
async def registrar_leitura(input: LeituraCreate, conferente_id: str):
    # Buscar carga
    carga = await db.cargas.find_one({"id": input.carga_id})
    if not carga:
        raise HTTPException(status_code=404, detail="Carga não encontrada")
    
    # Buscar sessão para pegar o recipiente (se Multi)
    sessao = await db.sessoes.find_one({"id": input.sessao_id})
    recipiente_sessao = sessao.get("recipiente") if sessao else None
    
    # Normalizar EAN
    ean_normalizado = normalizar_ean(input.ean)
    
    # 1. Buscar em produto_eans para mapear EAN → SKU
    produto_ean = await db.produto_eans.find_one({"ean": ean_normalizado, "ativo": True})
    
    if produto_ean:
        # EAN encontrado no sistema de múltiplos EANs
        sku = produto_ean["sku"]
        ean_busca = ean_normalizado
    else:
        # 2. Fallback: buscar no sistema antigo (produtos.ean)
        produto = await db.produtos.find_one({"ean": input.ean})
        if produto:
            sku = produto["codigo_produto"]
            ean_busca = normalizar_ean(input.ean)
        else:
            # EAN não existe em nenhum cadastro - Sobra
            await registrar_sobra(input.carga_id, input.sessao_id, input.ean, input.quantidade, recipiente_sessao)
            
            leitura = Leitura(
                sessao_id=input.sessao_id,
                carga_id=input.carga_id,
                conferente_id=conferente_id,
                ean=input.ean,
                quantidade=input.quantidade,
                resultado="fora_lista"
            )
            await db.leituras.insert_one(leitura.model_dump())
            return leitura
    
    # 3. Buscar item na carga pelo SKU + EAN
    # IMPORTANTE: Match por código_produto E ean (não só código)
    item_idx = None
    for idx, item in enumerate(carga["itens"]):
        # Normalizar EAN do item para comparação
        item_ean_norm = normalizar_ean(item.get("ean", ""))
        
        # Para Multi, considerar também o recipiente
        if carga["tipo"] == "multi":
            if (item["codigo_produto"] == sku and 
                item_ean_norm == ean_busca and 
                item.get("recipiente") == recipiente_sessao):
                item_idx = idx
                break
        else:
            if item["codigo_produto"] == sku and item_ean_norm == ean_busca:
                item_idx = idx
                break
    
    if item_idx is None:
        # SKU/EAN existe mas não está nesta carga/recipiente - Sobra
        descricao = produto_ean.get("descricao") if produto_ean else (produto.get("descricao") if 'produto' in locals() else None)
        await registrar_sobra(input.carga_id, input.sessao_id, input.ean, input.quantidade, recipiente_sessao, descricao)
        
        leitura = Leitura(
            sessao_id=input.sessao_id,
            carga_id=input.carga_id,
            conferente_id=conferente_id,
            ean=input.ean,
            quantidade=input.quantidade,
            resultado="fora_lista"
        )
        await db.leituras.insert_one(leitura.model_dump())
        return leitura
    
    # 4. Atualizar quantidade conferida (SEM multiplicação)
    nova_qtd_conferida = carga["itens"][item_idx]["quantidade_conferida"] + input.quantidade
    qtd_esperada = carga["itens"][item_idx]["quantidade"]
    
    resultado = "ok" if nova_qtd_conferida == qtd_esperada else "diferenca"
    status = "ok" if nova_qtd_conferida == qtd_esperada else "diferenca"
    
    carga["itens"][item_idx]["quantidade_conferida"] = nova_qtd_conferida
    carga["itens"][item_idx]["status"] = status
    
    await db.cargas.update_one(
        {"id": input.carga_id},
        {"$set": {"itens": carga["itens"]}}
    )
    
    leitura = Leitura(
        sessao_id=input.sessao_id,
        carga_id=input.carga_id,
        conferente_id=conferente_id,
        ean=input.ean,
        quantidade=input.quantidade,
        resultado=resultado
    )
    await db.leituras.insert_one(leitura.model_dump())
    return leitura

async def registrar_sobra(carga_id: str, sessao_id: str, ean: str, quantidade: int, recipiente: Optional[str] = None, descricao: Optional[str] = None):
    # Buscar se já existe sobra deste EAN
    filtro = {"carga_id": carga_id, "sessao_id": sessao_id, "ean": ean}
    if recipiente:
        filtro["recipiente"] = recipiente
    
    sobra_existente = await db.sobras.find_one(filtro)
    
    if sobra_existente:
        # Atualizar quantidade e timestamp
        await db.sobras.update_one(
            {"id": sobra_existente["id"]},
            {
                "$inc": {"quantidade": quantidade},
                "$set": {"ultima_leitura": datetime.now(timezone.utc).isoformat()}
            }
        )
    else:
        # Criar nova sobra
        sobra = Sobra(
            carga_id=carga_id,
            sessao_id=sessao_id,
            recipiente=recipiente,
            ean=ean,
            descricao=descricao,
            quantidade=quantidade
        )
        await db.sobras.insert_one(sobra.model_dump())

@api_router.get("/leituras/{sessao_id}", response_model=List[Leitura])
async def listar_leituras(sessao_id: str):
    leituras = await db.leituras.find({"sessao_id": sessao_id}, {"_id": 0}).to_list(10000)
    return leituras

# Sobras
@api_router.get("/sobras/{sessao_id}", response_model=List[Sobra])
async def listar_sobras(sessao_id: str):
    sobras = await db.sobras.find({"sessao_id": sessao_id}, {"_id": 0}).to_list(1000)
    return sobras

# Dashboard
@api_router.get("/dashboard/estatisticas")
async def obter_estatisticas(data: Optional[str] = None, tipo: Optional[str] = None, conferente_id: Optional[str] = None):
    filtro = {}
    if data:
        filtro["data"] = data
    if tipo:
        filtro["tipo"] = tipo
    if conferente_id:
        filtro["conferente_id"] = conferente_id
    
    cargas = await db.cargas.find(filtro, {"_id": 0}).to_list(1000)
    
    stats = []
    for carga in cargas:
        total_itens = len(carga["itens"])
        itens_ok = sum(1 for item in carga["itens"] if item["status"] == "ok")
        itens_diferenca = sum(1 for item in carga["itens"] if item["status"] == "diferenca")
        progresso = (itens_ok / total_itens * 100) if total_itens > 0 else 0
        
        # Buscar sessão
        sessao = await db.sessoes.find_one({"carga_id": carga["id"]}, {"_id": 0})
        tempo_total = None
        if sessao and sessao.get("fim"):
            inicio = datetime.fromisoformat(sessao["inicio"])
            fim = datetime.fromisoformat(sessao["fim"])
            tempo_total = int((fim - inicio).total_seconds() / 60)
        
        conferente_nome = "N/A"
        if carga.get("conferente_id"):
            conferente = await db.usuarios.find_one({"id": carga["conferente_id"]})
            if conferente:
                conferente_nome = conferente["nome"]
        
        stats.append({
            "identificador_carga": carga["identificador_carga"],
            "tipo": carga["tipo"],
            "status": carga["status"],
            "progresso": round(progresso, 2),
            "itens_ok": itens_ok,
            "itens_diferenca": itens_diferenca,
            "total_itens": total_itens,
            "conferente": conferente_nome,
            "tempo_minutos": tempo_total,
            "data": carga["data"]
        })
    
    return stats

# Admin
@api_router.post("/admin/limpar-dados")
async def limpar_dados():
    # Apagar apenas cargas finalizadas e leituras
    await db.cargas.delete_many({"status": "finalizada"})
    await db.leituras.delete_many({})
    await db.sessoes.delete_many({"status": "finalizada"})
    return {"message": "Dados locais limpos com sucesso"}

@api_router.post("/admin/resetar")
async def resetar_banco(senha_admin: str):
    if senha_admin != "945801":
        raise HTTPException(status_code=403, detail="Senha administrativa incorreta")
    
    # Apagar todas as collections exceto usuários
    await db.produtos.delete_many({})
    await db.cargas.delete_many({})
    await db.sessoes.delete_many({})
    await db.leituras.delete_many({})
    await db.produto_eans.delete_many({})
    
    return {"message": "Banco resetado com sucesso"}

@api_router.post("/admin/migrar-eans")
async def migrar_eans_produtos():
    """
    Migra EANs da tabela produtos para produto_eans.
    Cria um registro em produto_eans para cada produto existente.
    """
    produtos = await db.produtos.find({}, {"_id": 0}).to_list(10000)
    
    migrados = 0
    erros = []
    
    for produto in produtos:
        if not produto.get("ean"):
            continue
            
        ean_normalizado = normalizar_ean(produto["ean"])
        
        # Verificar se já existe
        existe = await db.produto_eans.find_one({"ean": ean_normalizado})
        if existe:
            erros.append(f"EAN {produto['ean']} já existe em produto_eans")
            continue
        
        # Criar registro
        produto_ean = ProdutoEAN(
            sku=produto["codigo_produto"],
            ean=ean_normalizado,
            tipo_unidade=produto.get("tipo_unidade", "UNI"),
            fator_conversao=1,  # Padrão: 1 unidade
            descricao=produto.get("descricao"),
            ativo=produto.get("ativo", True)
        )
        
        await db.produto_eans.insert_one(produto_ean.model_dump())
        migrados += 1
    
    return {
        "message": f"Migração concluída",
        "migrados": migrados,
        "total_produtos": len(produtos),
        "erros": erros
    }

# Inicialização
@app.on_event("startup")
async def startup_event():
    # Criar usuário admin padrão se não existir
    admin_existe = await db.usuarios.find_one({"login": "admin"})
    if not admin_existe:
        admin = Usuario(
            nome="Administrador",
            login="admin",
            senha_hash=hash_senha("admin123"),
            perfil="gestor"
        )
        await db.usuarios.insert_one(admin.model_dump())
        logging.info("Usuário admin criado: login=admin, senha=admin123")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
